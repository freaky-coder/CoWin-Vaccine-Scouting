# -*- coding: utf-8 -*-
"""
Created on Thu May 5 10:38:32 2021
@author: Ayush Kapoor
"""
#%% Importing Necessary Libraries
import requests
import time
from datetime import datetime
from pytz import timezone
import json
import smtplib,ssl
from json2html import *
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import load_workbook
import urllib3

#%% Variable Declarations
file_path = "C:\\Users\\ayush.kapoor\\Desktop\\input.xlsx" 
wb = load_workbook(filename= file_path,data_only=True)
sheet = wb['Details']

#%% Function Declarations
# Function to get data from Google sheets and store it in a local database
def getfromGSheets():
    scope = ["https://spreadsheets.google.com/feeds",'https://www.googleapis.com/auth/spreadsheets',"https://www.googleapis.com/auth/drive.file","https://www.googleapis.com/auth/drive"]     
    creds = ServiceAccountCredentials.from_json_keyfile_name("drive_api.json", scope)
    wb = load_workbook(filename= file_path,data_only=True)
    sheetOffline = wb['Details']
    client = gspread.authorize(creds)
    Gsheet = client.open("Vaccine Scouting 18-44 Age (Responses)").sheet1
    #row = sheet.row_values(3)  # Get a specific row
    #col = sheet.col_values(3)  # Get a specific column
    i=2
    while (Gsheet.cell(i,4)!=None):
        sheetOffline.cell(i,1).value = Gsheet.cell(i,5).value
        sheetOffline.cell(i,2).value = Gsheet.cell(i,6).value
        sheetOffline.cell(i,3).value = Gsheet.cell(i,7).value
        sheetOffline.cell(i,4).value = Gsheet.cell(i,8).value
        i+=1
    wb.save(filename=file_path) 

# Function to get all the states and districts on the CoWin portal    
def fetchStateDistrict():
    wb = load_workbook(filename= file_path,data_only=True)
    sheet = wb['Database']
    k = 1
    cowinS = "https://api.cowin.gov.in/api/v2/admin/location/states/"
    stateResp = json.loads(requests.get(cowinS,verify=False).text)
    for j in range(0,len(stateResp['states'])):
        cowin_state_url = "https://api.cowin.gov.in/api/v2/admin/location/districts/"+ str(stateResp['states'][j]['state_id']) # 1 to 37 for district
        print(cowin_state_url)
        distReq = requests.get(cowin_state_url,verify=False)
        distResp = json.loads(distReq.text)                        
        for i in range(0,len(distResp['districts'])):
            k+=1
            sheet.cell(row=k,column=3).value = distResp['districts'][i]['district_name']
            sheet.cell(row=k,column=4).value = distResp['districts'][i]['district_id']
            sheet.cell(row=k,column=2).value = stateResp['states'][j]['state_id']
            sheet.cell(row=k,column=1).value = stateResp['states'][j]['state_name']
    wb.save(filename=file_path)
    print("Database Fetched")
    
# A helper function to make the email text html friendly    
def HTMLReady (str_ , name):
    if(str_ == "opners"):
        return "<!DOCTYPE html><html><body font-family: Arial;>"
    if(str_ == "mail_header"):
        return "<p>Hi "+name+','+"<br><br>We noticed the availability of COVID-19 vaccines at vaccination centres as per your registered location. Please find below the requisite details."
    if(str_ == "mail_footer"):
        return "<p><br>Thanks,</p>"
    if(str_ == "closers"):
        return "</body></html>"
    if(str_ == "getCss"):
        return "<style>body {  font-family: Arial;}</style>"
    return ""

# Function to convert the JSON received from the API to HTML        
def JSON2HTML(listObj , name):
    htmlTable =""
    htmlTable = HTMLReady("openers","")
    htmlTable+= HTMLReady("mail_header",name)
    obj = Json2Html()
    for i in range(0,len(listObj)):
        temp = json.loads(json.dumps(listObj[i]))
        del temp['session_id']
        temp['slots'] = str(temp['slots']).replace(":00","").replace("'",'').replace("0",'').replace('[','').replace(']','')
        del temp['center']['center_id']
        temp['center']['center_name'] = temp['center']['center_name']+', '+temp['center']['block_name']+', '+str(temp['center']['pincode'])
        del temp['center']['block_name']
        del temp['center']['pincode']
        temp['center'] = temp['center']['center_name']+' ['+temp['center']['fee_type']+']'
        htmlTable+= obj.convert(json.dumps(temp))
        htmlTable+="<br>"
    htmlTable = htmlTable.replace("<ul>","").replace("</ul>","").replace("<li>"," ").replace("</li>","")
    htmlTable = htmlTable.replace("date","Date").replace("available_capacity", "Available Slots").replace("min_age_limit","Minimum Age").replace('Available Slots_dose1', 'Available Slots (Dose 1)').replace('Available Slots_dose2', 'Available Slots (Dose 2)')   
    htmlTable = htmlTable.replace("vaccine","Vaccine").replace("slots","Slot Timings")
    htmlTable = htmlTable.replace("center","Centre Description")
    htmlTable+= HTMLReady("getCss","")
    htmlTable+= HTMLReady("mail_footer","")
    htmlTable+= '<table cellpadding="0" cellspacing="0" class="sc-gPEVay eQYmiW" style="vertical-align: -webkit-baseline-middle; font-size: small; font-family: Arial;"><tbody><tr><td><table cellpadding="0" cellspacing="0" class="sc-gPEVay eQYmiW" style="vertical-align: -webkit-baseline-middle; font-size: small; font-family: Arial;"><tbody><tr><td style="vertical-align: top;"><h3 color="#000000" class="sc-fBuWsC eeihxG" style="margin: 0px; font-size: 16px; color: rgb(0, 0, 0);"><span>Ayush</span><span>&nbsp;</span><span>Kapoor</span></h3><table cellpadding="0" cellspacing="0" class="sc-gPEVay eQYmiW" style="vertical-align: -webkit-baseline-middle; font-size: small; font-family: Arial; width: 100%;"><tbody><tr><td height="4"></td></tr><tr><td color="#7389f6" direction="horizontal" height="1" class="sc-jhAzac hmXDXQ" style="width: 100%; border-bottom: 1px solid rgb(115, 137, 246); border-left: none; display: block;"></td></tr><tr><td height="10"></td></tr></tbody></table><table cellpadding="0" cellspacing="0" class="sc-gPEVay eQYmiW" style="vertical-align: -webkit-baseline-middle; font-size: small; font-family: Arial;"><tbody><tr height="20" style="vertical-align: middle;"><td width="30" style="vertical-align: middle;"><table cellpadding="0" cellspacing="0" class="sc-gPEVay eQYmiW" style="vertical-align: -webkit-baseline-middle; font-size: small; font-family: Arial;"><tbody><tr><td style="vertical-align: bottom;"><span color="#7389f6" width="11" class="sc-jlyJG bbyJzT" style="display: block; background-color: rgb(115, 137, 246);"><img src="https://cdn2.hubspot.net/hubfs/53/tools/email-signature-generator/icons/email-icon-2x.png" color="#7389f6" width="13" class="sc-iRbamj blSEcj" style="display: block; background-color: rgb(115, 137, 246);"></span></td></tr></tbody></table></td><td style="padding: 0px;"><a href="mailto:ak246@snu.edu.in" color="#000000" class="sc-gipzik iyhjGb" style="text-decoration: none; color: rgb(0, 0, 0); font-size: 12px;"><span>ak246@snu.edu.in</span></a></td></tr><tr height="25" style="vertical-align: middle;"><td width="30" style="vertical-align: middle;"><table cellpadding="0" cellspacing="0" class="sc-gPEVay eQYmiW" style="vertical-align: -webkit-baseline-middle; font-size: small; font-family: Arial;"><tbody><tr><td style="vertical-align: bottom;"><span color="#7389f6" width="11" class="sc-jlyJG bbyJzT" style="display: block; background-color: rgb(115, 137, 246);"><img src="https://cdn2.hubspot.net/hubfs/53/tools/email-signature-generator/icons/address-icon-2x.png" color="#7389f6" width="13" class="sc-iRbamj blSEcj" style="display: block; background-color: rgb(115, 137, 246);"></span></td></tr></tbody></table></td><td style="padding: 0px;"><span color="#000000" class="sc-csuQGl CQhxV" style="font-size: 12px; color: rgb(0, 0, 0);"><span>IN</span></span></td></tr></tbody></table></td><td width="40"><div></div></td><td style="vertical-align: top;"><table cellpadding="0" cellspacing="0" class="sc-gPEVay eQYmiW" style="vertical-align: -webkit-baseline-middle; font-size: small; font-family: Arial;"><tbody><tr><td style="text-align: center;"><table cellpadding="0" cellspacing="0" class="sc-gPEVay eQYmiW" style="vertical-align: -webkit-baseline-middle; font-size: small; font-family: Arial; display: inline-block;"><tbody><tr style="text-align: center;"><td><a href="https://www.linkedin.com/in/kapoor-ayush/" color="#4451dd" class="sc-hzDkRC kpsoyz" style="display: inline-block; padding: 0px; background-color: rgb(68, 81, 221);"><img src="https://cdn2.hubspot.net/hubfs/53/tools/email-signature-generator/icons/linkedin-icon-2x.png" alt="linkedin" color="#4451dd" height="24" class="sc-bRBYWo ccSRck" style="background-color: rgb(68, 81, 221); max-width: 135px; display: block;"></a></td><td width="5"><div></div></td></tr></tbody></table></td></tr></tbody></table></td></tr></tbody></table></td></tr><tr><td height="30"></td></tr></tbody></table>'
    htmlTable+= HTMLReady("closers","")
    return htmlTable 

# The most crucial function of the entire script that fetches data from the CoWin API
def fetchDetails(apiURL,districtId,numRows):
    # Conditions of fetching from the CoWin API
    min_age = 45
    capacity = 1
    local_timezone = timezone("Asia/Kolkata")
    local_date = local_timezone.localize(datetime.now())
    api_params = {"district_id": districtId, "date": local_date.strftime("%d-%m-%Y")}
    api_request = requests.get(apiURL,params=api_params,verify = False)
    api_response = json.loads(api_request.text)
    centers = []
    for center in api_response["centers"]:
        centers.append(center["name"])
    if (len(centers)!= 0):
        length = len(max(centers, key=len))
    sessionData = []
    for center in api_response["centers"]:
        for session in center["sessions"]:
            # or (session["min_age_limit"]<min_age and session["available_capacity_dose2"]>capacity)
            if((session["min_age_limit"]<min_age and session["available_capacity_dose1"]>capacity)):
                session["center"]={"center_id":center["center_id"],"center_name":center["name"],"block_name":center["block_name"],"pincode":center["pincode"],"fee_type":center["fee_type"]}
                print(length,center["name"],session["date"],"SUCCESS")
                sessionData.append(session)
    print('Total 18+ Centres : '+str(len(sessionData)))
    if(len(sessionData)>0):
        sendMail(districtId,sessionData,numRows)
        
# Function to send mail to the respective users
def sendMail(districtId,data,numRows):
    # Server Credentials
    HOSTNAME = 'smtp.gmail.com'  
    PORT = '465'
    CONTEXT= ssl.create_default_context()
    from_password= "Ch3cooh!@"     
    i=2
    while(i<=numRows):
        if ((sheet.cell(i,7).value)==districtId and sheet.cell(i,1).value == "No"):
            print(str(sheet.cell(i,7).value))
            print(sheet.cell(i,3).value)
            msg = MIMEMultipart('alternative')
            email_message = ""
            msg['From'] = "Ayush Kapoor ayushmailer111@gmail.com"
            msg['Subject'] = "ATTENTION: CoWin Vaccination slots for 18+ available at "+ str(len(data))+" centres in "+sheet.cell(i,6).value+', '+sheet.cell(i,5).value
            msg['To'] = sheet.cell(i,4).value
            print(sheet.cell(i,4).value)
            email_message= JSON2HTML(data, sheet.cell(i,3).value)
            part1 = MIMEText(email_message, 'html')
            msg.attach(part1)
            with smtplib.SMTP_SSL(host=HOSTNAME,port=PORT,context=CONTEXT) as server:
                server.login(msg['From'].split(' ')[2],from_password)
                server.sendmail(msg['From'],msg['To'],msg.as_string())
        i+=1 
    
#%% 
#fetchStateDistrict()
#key = ""
#%% Main 
# Driver setup
dataDict = {}
tempList = [] 
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
cowin_api_url2 = "https://api.cowin.gov.in/api/v2/appointment/sessions/public/calendarByDistrict"
cowin_api_url = "https://cdn-api.co-vin.in/api/v2/appointment/sessions/public/findByDistrict"
districtCodes = []
#while True:
i=2
count = 0
while(i<=sheet.cell(2,9).value):
    if (sheet.cell(i,1).value =="No" and sheet.cell(i,6).value!=""):
        print ('* '+sheet.cell(i,3).value+' | '+ str(sheet.cell(i,7).value))
        dataDict[sheet.cell(i,7).value] = (sheet.cell(i,3).value)
        districtCodes.append(sheet.cell(i,7).value)
        print (tempList)
    del tempList[:]
    count+=1
    i+=1
numRows = i-1
districtCodes = list(set(districtCodes))
#%%
print('-----Last Row: '+str(numRows))
for j in range(0,len(districtCodes)):
    print('------>'+str(districtCodes[j]))
    fetchDetails(cowin_api_url2,districtCodes[j],numRows)
print('--------Last Fetched: '+datetime.now().strftime("%d/%m/%Y %H:%M:%S")+'------')
